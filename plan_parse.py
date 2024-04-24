import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from classes import ControlForm, TotalHours, RequiredHours, Semester, Discipline


class Plan:
    """
    Класс, хранящий в себе всю информацию из одного файла учебного плана.
    Имеет методы для обработки этого файла и их сохранения в объекты класса
    """

    # file_path - путь к файлу учебного плана
    def __init__(self, file_path: str):
        # Открываем файл и получаем объект листа с планом
        self.plan_worksheet = self.__open_worksheet_in_file(file_path)
        # Получение информации с титульного листа документа
        self.get_title_info()
        # Получение информации про все дисциплины из файла
        self.disciplines: list[Discipline] = self.get_disciplines()

    def __open_worksheet_in_file(self, file_path: str) -> Worksheet:
        """Открывает файл по его пути и возвращает лист с планом"""
        # Загрузка файла Excel
        self.workbook: Workbook = openpyxl.load_workbook(file_path)
        # Получение листа по имени
        return self.workbook['План']

    @staticmethod
    def __is_bold(cell: Cell) -> bool:
        """Проверка ячейки на жирность шрифта"""
        return cell.font.bold

    @staticmethod
    def __check_value_and_split(value: str | None) -> list[int] | None:
        """Проверка значения на None и разделение его посимвольно, с превращением в числа"""
        if value:
            return [int(x) for x in value]
        else:
            return None

    @staticmethod
    def __check_int(value: str | None) -> int | None:
        """Проверка значения на None и превращение в число"""
        if value:
            return int(value)
        else:
            return None

    def __get_control_form(self, row: list[str], sem_num: int) -> ControlForm:
        """Получение информации о том, какая именно форма контроля у дисциплины в переданной строке"""
        sem = None
        col = 3
        while sem is None and col < 9:
            sem = self.__check_value_and_split(row[col])
            col += 1

        if sem is not None and sem_num in sem:
            match col:
                case 3:
                    return ControlForm.EKZ
                case 4:
                    return ControlForm.ZACH
                case 5:
                    return ControlForm.ZACH0
                case 6:
                    return ControlForm.KP
                case 7:
                    return ControlForm.KR
                case 8:
                    return ControlForm.DR
                case _:
                    return ControlForm.NO
        else:
            return ControlForm.NO

    def __get_semesters_from_row(self, row: list[str]) -> list[Semester]:
        """Получение информации про семестры и часы в переданной строке"""
        semesters: list[Semester] = []
        sem_row = row[17:]

        for i in range(0, len(sem_row) - 2, 9):
            if sem_row[i]:
                num = i // 9 + 1
                total = self.__check_int(sem_row[i])
                lek = self.__check_int(sem_row[i + 1])
                lab = self.__check_int(sem_row[i + 2])
                pr = self.__check_int(sem_row[i + 3])
                krp = self.__check_int(sem_row[i + 4])
                ip = self.__check_int(sem_row[i + 5])
                sr = self.__check_int(sem_row[i + 6])
                cons = self.__check_int(sem_row[i + 7])
                patt = self.__check_int(sem_row[i + 8])
                control = self.__get_control_form(row, num)
                semesters.append(Semester(num, total, lek, lab, pr, krp, ip, sr, cons, patt, control))

        return semesters

    def get_total_hours(self, row: list[str]) -> TotalHours:
        """Получение общего количества академических часов в переданной строке"""
        expert = self.__check_int(row[9])
        plan = self.__check_int(row[10])
        with_teacher = self.__check_int(row[11])
        ip = self.__check_int(row[12])
        sr = self.__check_int(row[13])
        patt = self.__check_int(row[14])

        return TotalHours(expert, plan, with_teacher, ip, sr, patt)

    def get_required(self, row: list[str]) -> RequiredHours:
        """Получение объёма ОП в переданной строке"""
        important = self.__check_int(row[15])
        not_important = self.__check_int(row[16])

        return RequiredHours(important, not_important)

    def get_disciplines(self) -> list[Discipline]:
        """Получение всех дисциплин в файле и информации про них"""
        row: tuple[Cell]
        disciplines: list[Discipline] = []
        for row in self.plan_worksheet.iter_rows(min_row=6, max_row=self.plan_worksheet.max_row):
            if self.__is_bold(row[2]) or row[2].value is None:
                continue
            row: list[str] = [cell.value for cell in row]

            # Получение общего кол-ва часов
            total_hours = self.get_total_hours(row)
            # Получение обязательной программы
            required = self.get_required(row)
            # Получение семестров, в которых будут формы контроля
            semesters = self.__get_semesters_from_row(row)

            # Получение индекса, названия и статуса дисциплины
            in_plan = True if row[0] == '+' else False
            ind = row[1]
            name = row[2]
            disciplines.append(Discipline(in_plan, ind, name, total_hours, required, semesters))

        return disciplines

    def get_title_info(self):
        """Получение информации с титульного листа в файле"""
        title_worksheet = self.workbook['Титул']
        cell_value_list = title_worksheet['D29'].value.split()
        if '_x000d_' in cell_value_list:
            cell_value_list.remove('_x000d_')
        ind = cell_value_list.index('Профиль')
        self.name = ' '.join(cell_value_list[:ind])
        self.cafedra = title_worksheet['D37'].value
        self.facultet = title_worksheet['D38'].value
        self.profile = ' '.join(cell_value_list[ind + 10:])
        self.cod = title_worksheet['D27'].value
        self.kvalik = title_worksheet['C40'].value.split(':')[1]
        self.edu_form = title_worksheet['C42'].value.split(':')[1]
        self.start_year = int(title_worksheet['W40'].value) if title_worksheet['W40'].value else 0
        self.standart = title_worksheet['W42'].value
        self.baza = title_worksheet['C44'].value.split(':')[1]
